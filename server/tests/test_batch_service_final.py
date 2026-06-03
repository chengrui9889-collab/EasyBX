from datetime import date
from io import BytesIO

import pytest
from fastapi import HTTPException

from app.schemas.batch import AddInvoicesRequest, UpdateBatchRequest


def _make_user(db, username):
    from app.models.user import User
    user = User(username=username, password_hash="hash")
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def _make_invoice(db, user_id, status="confirmed", amount=100.0, **kwargs):
    from app.models.invoice import Invoice
    inv = Invoice(
        user_id=user_id,
        file_path=f"/tmp/{user_id}/test.pdf",
        status=status,
        amount=amount,
        **kwargs,
    )
    db.add(inv)
    db.commit()
    db.refresh(inv)
    return inv


def _make_batch(db, user_id, department="测试", reporter="测试", **kwargs):
    from app.services.batch_service import create_batch
    from app.schemas.batch import CreateBatchRequest
    return create_batch(db, user_id, CreateBatchRequest(department=department, reporter=reporter, **kwargs))


def _add_invoice_to_batch(db, user_id, batch_id, invoice_id):
    from app.services.batch_service import add_invoices
    return add_invoices(db, user_id, batch_id, AddInvoicesRequest(invoice_ids=[invoice_id]))


class TestUpdateBatch:
    def test_update_department(self, db):
        from app.services.batch_service import update_batch

        user = _make_user(db, "upd_batch1")
        batch = _make_batch(db, user.id, "测试", "张三")

        result = update_batch(db, user.id, batch.id, UpdateBatchRequest(department="新部门"))
        assert result.department == "新部门"

    def test_update_reviewer(self, db):
        from app.services.batch_service import update_batch

        user = _make_user(db, "upd_batch2")
        batch = _make_batch(db, user.id)

        result = update_batch(db, user.id, batch.id, UpdateBatchRequest(reviewer="张经理"))
        assert result.reviewer == "张经理"

    def test_update_partial_fields_preserves_others(self, db):
        from app.services.batch_service import update_batch

        user = _make_user(db, "upd_batch3")
        batch = _make_batch(db, user.id, "研发", "张三")

        result = update_batch(db, user.id, batch.id, UpdateBatchRequest(reviewer="李经理"))
        assert result.department == "研发"
        assert result.reviewer == "李经理"

    def test_batch_not_found(self, db):
        from app.services.batch_service import update_batch

        user = _make_user(db, "upd_batch4")
        with pytest.raises(HTTPException) as exc:
            update_batch(db, user.id, 99999, UpdateBatchRequest(department="X"))
        assert exc.value.status_code == 404

    def test_wrong_user(self, db):
        from app.services.batch_service import update_batch

        user1 = _make_user(db, "upd_batch5a")
        user2 = _make_user(db, "upd_batch5b")
        batch = _make_batch(db, user1.id)

        with pytest.raises(HTTPException) as exc:
            update_batch(db, user2.id, batch.id, UpdateBatchRequest(department="X"))
        assert exc.value.status_code == 404


class TestDeleteBatch:
    def test_delete_with_invoices(self, db):
        from app.services.batch_service import delete_batch

        user = _make_user(db, "del_batch1")
        inv1 = _make_invoice(db, user.id, status="confirmed", amount=100)
        inv2 = _make_invoice(db, user.id, status="confirmed", amount=200)
        inv3 = _make_invoice(db, user.id, status="confirmed", amount=300)
        inv4 = _make_invoice(db, user.id, status="confirmed", amount=400)
        inv5 = _make_invoice(db, user.id, status="confirmed", amount=500)
        batch = _make_batch(db, user.id)
        from app.services.batch_service import add_invoices
        add_invoices(db, user.id, batch.id, AddInvoicesRequest(
            invoice_ids=[inv1.id, inv2.id, inv3.id, inv4.id, inv5.id]
        ))

        result = delete_batch(db, user.id, batch.id)
        assert result["deleted"] is True
        assert result["released_invoice_count"] == 5

    def test_deleted_invoices_reappear_in_available(self, db):
        from app.services.batch_service import delete_batch, list_available_invoices

        user = _make_user(db, "del_batch2")
        inv = _make_invoice(db, user.id, status="confirmed", amount=100)
        batch = _make_batch(db, user.id)
        _add_invoice_to_batch(db, user.id, batch.id, inv.id)

        delete_batch(db, user.id, batch.id)

        avail = list_available_invoices(db, user.id)
        assert avail.total == 1
        assert avail.items[0].id == inv.id

    def test_delete_empty_batch(self, db):
        from app.services.batch_service import delete_batch

        user = _make_user(db, "del_batch3")
        batch = _make_batch(db, user.id)

        result = delete_batch(db, user.id, batch.id)
        assert result["deleted"] is True
        assert result["released_invoice_count"] == 0

    def test_batch_not_found(self, db):
        from app.services.batch_service import delete_batch

        user = _make_user(db, "del_batch4")
        with pytest.raises(HTTPException) as exc:
            delete_batch(db, user.id, 99999)
        assert exc.value.status_code == 404

    def test_delete_cascades_substitute_relations(self, db):
        from app.services.batch_service import delete_batch
        from app.models.batch import BatchInvoice
        from app.models.substitute import SubstituteRelation

        user = _make_user(db, "del_batch5")
        sub_inv = _make_invoice(db, user.id, status="confirmed", amount=500.0, invoice_no="SUB001")
        batch = _make_batch(db, user.id)

        manual = BatchInvoice(
            batch_id=batch.id,
            source_type="manual",
            expense_item="奖金",
            row_amount=500.0,
        )
        db.add(manual)
        db.commit()
        db.refresh(manual)

        placeholder = BatchInvoice(
            batch_id=batch.id,
            invoice_id=sub_inv.id,
            source_type="invoice",
            is_substitute=True,
        )
        db.add(placeholder)
        db.commit()

        rel = SubstituteRelation(
            batch_id=batch.id,
            substitute_invoice_id=sub_inv.id,
            target_row_id=manual.id,
            mode="one_to_one",
        )
        db.add(rel)
        db.commit()

        assert db.query(SubstituteRelation).filter(SubstituteRelation.batch_id == batch.id).count() == 1

        delete_batch(db, user.id, batch.id)

        assert db.query(SubstituteRelation).filter(SubstituteRelation.batch_id == batch.id).count() == 0


class TestExportBatchExcel:
    def test_export_no_invoices_returns_400(self, db):
        from app.services.excel_service import export_batch_excel

        user = _make_user(db, "exp_batch1")
        batch = _make_batch(db, user.id)

        with pytest.raises(HTTPException) as exc:
            export_batch_excel(db, user.id, batch.id)
        assert exc.value.status_code == 400

    def test_export_with_invoices_returns_bytes(self, db):
        from app.services.excel_service import export_batch_excel

        user = _make_user(db, "exp_batch2")
        inv1 = _make_invoice(db, user.id, status="confirmed", amount=100.0,
                             invoice_date=date(2026, 5, 15), category="交通",
                             vendor="铁路总公司")
        inv2 = _make_invoice(db, user.id, status="confirmed", amount=200.0,
                             invoice_date=date(2026, 5, 16), category="住宿",
                             vendor="某酒店")
        inv3 = _make_invoice(db, user.id, status="confirmed", amount=50.0,
                             invoice_date=date(2026, 5, 17), category="餐饮")
        batch = _make_batch(db, user.id, "研发部", "张三",
                            period_start=date(2026, 5, 1), period_end=date(2026, 5, 31),
                            report_date=date(2026, 5, 20), reviewer="李经理",
                            payee="张三", bank_account="6222000012345678", bank_name="工商银行")

        from app.services.batch_service import add_invoices
        add_invoices(db, user.id, batch.id, AddInvoicesRequest(
            invoice_ids=[inv1.id, inv2.id, inv3.id]
        ))

        result = export_batch_excel(db, user.id, batch.id)
        assert isinstance(result, bytes)
        assert len(result) > 0

        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Sheet1"]

        assert "报账部门：研发部" in str(ws["A2"].value)
        assert "报账期间：2026-05-01-2026-05-31" in str(ws["A2"].value)

        assert ws["A4"].value is not None
        assert ws["A5"].value is not None
        assert ws["A6"].value is not None

        assert "=SUM(F4:F6)" in str(ws["F7"].value)

        assert "审核人：李经理" in str(ws["A8"].value or "")
        assert "报账人：张三" in str(ws["A8"].value or "")
        assert "收款人：张三" in str(ws["A9"].value or "")
        assert "银行卡号：6222000012345678" in str(ws["A9"].value or "")
        assert "开户行：工商银行" in str(ws["A9"].value or "")

    def test_export_zero_amount_invoice(self, db):
        from app.services.excel_service import export_batch_excel

        user = _make_user(db, "exp_batch3")
        inv = _make_invoice(db, user.id, status="confirmed", amount=0.0,
                            invoice_date=date(2026, 5, 15))
        batch = _make_batch(db, user.id)
        _add_invoice_to_batch(db, user.id, batch.id, inv.id)

        result = export_batch_excel(db, user.id, batch.id)
        assert isinstance(result, bytes)
        assert len(result) > 0

        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Sheet1"]
        assert ws["E4"].value == 0.0
        assert ws["A4"].value is not None

    def test_batch_not_found(self, db):
        from app.services.excel_service import export_batch_excel

        user = _make_user(db, "exp_batch4")
        with pytest.raises(HTTPException) as exc:
            export_batch_excel(db, user.id, 99999)
        assert exc.value.status_code == 404

    def test_export_with_manual_rows_only(self, db):
        from app.services.excel_service import export_batch_excel
        from app.models.batch import BatchInvoice

        user = _make_user(db, "exp_mr1")
        batch = _make_batch(db, user.id, "研发部", "张三",
                            period_start=date(2026, 5, 1), period_end=date(2026, 5, 31),
                            report_date=date(2026, 5, 20), reviewer="李经理")

        row1 = BatchInvoice(
            batch_id=batch.id,
            source_type="manual",
            expense_item="奖金",
            row_amount=1000.0,
            row_date=date(2026, 5, 15),
        )
        row2 = BatchInvoice(
            batch_id=batch.id,
            source_type="manual",
            expense_item="团建费",
            row_amount=500.0,
            row_date=date(2026, 5, 16),
        )
        db.add_all([row1, row2])
        db.commit()

        result = export_batch_excel(db, user.id, batch.id)
        assert isinstance(result, bytes)
        assert len(result) > 0

        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Sheet1"]

        assert ws["A4"].value == "2026-05-15"
        assert ws["B4"].value == "奖金"
        assert ws["E4"].value == 1000.0
        assert ws["A5"].value == "2026-05-16"
        assert ws["B5"].value == "团建费"
        assert ws["E5"].value == 500.0

        assert "报账人：张三" in str(ws["A7"].value or "")

    def test_export_excludes_substitute_placeholders(self, db):
        from app.services.excel_service import export_batch_excel
        from app.models.batch import BatchInvoice

        user = _make_user(db, "exp_mr2")
        inv1 = _make_invoice(db, user.id, status="confirmed", amount=300.0,
                             invoice_date=date(2026, 5, 15), category="交通")
        sub_inv = _make_invoice(db, user.id, status="confirmed", amount=500.0,
                                invoice_no="SUB001")
        batch = _make_batch(db, user.id)
        _add_invoice_to_batch(db, user.id, batch.id, inv1.id)

        placeholder = BatchInvoice(
            batch_id=batch.id,
            invoice_id=sub_inv.id,
            source_type="invoice",
            is_substitute=True,
        )
        db.add(placeholder)
        db.commit()

        result = export_batch_excel(db, user.id, batch.id)
        assert isinstance(result, bytes)

        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Sheet1"]

        assert ws["B4"].value == "交通"

    def test_export_mixed_invoices_and_manual_rows(self, db):
        from app.services.excel_service import export_batch_excel
        from app.models.batch import BatchInvoice

        user = _make_user(db, "exp_mr3")
        inv = _make_invoice(db, user.id, status="confirmed", amount=200.0,
                            invoice_date=date(2026, 5, 15), category="交通")
        batch = _make_batch(db, user.id)
        _add_invoice_to_batch(db, user.id, batch.id, inv.id)

        manual = BatchInvoice(
            batch_id=batch.id,
            source_type="manual",
            expense_item="奖金",
            row_amount=300.0,
            row_date=date(2026, 5, 16),
        )
        db.add(manual)
        db.commit()

        result = export_batch_excel(db, user.id, batch.id)
        assert isinstance(result, bytes)

        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Sheet1"]

        categories = {ws[f"B{i}"].value for i in range(4, 7) if ws[f"B{i}"].value}
        assert "交通" in categories
        assert "奖金" in categories

    def test_export_manual_row_advance_amount(self, db):
        from app.services.excel_service import export_batch_excel
        from app.models.batch import BatchInvoice

        user = _make_user(db, "exp_adv1")
        batch = _make_batch(db, user.id, "研发部", "张三")
        manual = BatchInvoice(
            batch_id=batch.id,
            source_type="manual",
            expense_item="差旅费",
            row_amount=1000.0,
            advance_amount=800.0,
            row_date=date(2026, 5, 15),
        )
        db.add(manual)
        db.commit()

        result = export_batch_excel(db, user.id, batch.id)
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Sheet1"]

        assert ws["F4"].value == 800.0, f"Expected 800.0, got {ws['F4'].value}"

    def test_export_substituted_manual_row_advance_amount(self, db):
        from app.services.excel_service import export_batch_excel
        from app.models.batch import BatchInvoice

        user = _make_user(db, "exp_adv2")
        batch = _make_batch(db, user.id, "研发部", "张三")
        manual = BatchInvoice(
            batch_id=batch.id,
            source_type="manual",
            expense_item="差旅费",
            row_amount=1000.0,
            advance_amount=800.0,
            row_date=date(2026, 5, 15),
            is_substitute=True,
            substitute_for="替票SUB001",
        )
        db.add(manual)
        db.commit()

        result = export_batch_excel(db, user.id, batch.id)
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(result))
        ws = wb["Sheet1"]

        assert ws["F4"].value == 800.0, f"Expected 800.0, got {ws['F4'].value}"