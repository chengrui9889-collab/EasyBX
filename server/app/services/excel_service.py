from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
from fastapi import HTTPException
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.models.batch import BatchInvoice, ReimbursementBatch
from app.models.invoice import Invoice

TEMPLATE_PATH = Path(__file__).parent.parent.parent.parent / "文件" / "台账模版.xlsx"


def export_batch_excel(db: Session, user_id: int, batch_id: int) -> bytes:
    batch = db.query(ReimbursementBatch).filter(
        ReimbursementBatch.id == batch_id,
        ReimbursementBatch.user_id == user_id,
    ).first()
    if not batch:
        raise HTTPException(status_code=404, detail="批次不存在")

    batch_invoices = (
        db.query(BatchInvoice, Invoice)
        .outerjoin(Invoice, BatchInvoice.invoice_id == Invoice.id)
        .filter(
            BatchInvoice.batch_id == batch_id,
            or_(
                BatchInvoice.source_type != "invoice",
                BatchInvoice.is_substitute == False,
                BatchInvoice.substitute_for.isnot(None),
            ),
        )
        .all()
    )

    if not batch_invoices:
        raise HTTPException(status_code=400, detail="请先添加发票")

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Sheet1"]

    if "Sheet3" in wb.sheetnames:
        del wb["Sheet3"]

    period_start = batch.period_start.isoformat() if batch.period_start else ""
    period_end = batch.period_end.isoformat() if batch.period_end else ""
    ws["A2"] = f"报账部门：{batch.department}  报账期间：{period_start}-{period_end}"

    row = 4
    for bi, inv in batch_invoices:
        if bi.source_type == "manual":
            ws[f"A{row}"] = bi.row_date.isoformat() if bi.row_date else ""
            ws[f"B{row}"] = bi.expense_item or ""
            ws[f"C{row}"] = 1
            ws[f"D{row}"] = bi.row_amount or 0.0
            ws[f"E{row}"] = bi.row_amount or 0.0
            ws[f"F{row}"] = bi.advance_amount or 0.0
            ws[f"G{row}"] = bi.remark or ""
        else:
            ws[f"A{row}"] = inv.invoice_date.isoformat() if inv and inv.invoice_date else ""
            ws[f"B{row}"] = inv.category if inv else ""
            ws[f"C{row}"] = bi.quantity
            ws[f"D{row}"] = bi.unit_price
            ws[f"E{row}"] = inv.amount if inv else 0.0
            ws[f"F{row}"] = bi.advance_amount
            ws[f"G{row}"] = bi.remark or ""
        row += 1

    last_data_row = row - 1

    if row <= 19:
        ws.delete_rows(row, 19 - row + 1)
        sum_row = row
        footer1_row = row + 1
        footer2_row = row + 2
    else:
        sum_row = row
        footer1_row = row + 1
        footer2_row = row + 2

    ws[f"F{sum_row}"] = f"=SUM(F4:F{last_data_row})"

    total_amount = 0.0
    for bi, inv in batch_invoices:
        if bi.source_type == "manual":
            total_amount += bi.row_amount or 0.0
        else:
            total_amount += inv.amount if inv else 0.0

    reviewer = batch.reviewer or ""
    reporter = batch.reporter or ""
    report_date = batch.report_date.isoformat() if batch.report_date else ""
    payee = batch.payee or ""
    bank_account = batch.bank_account or ""
    bank_name = batch.bank_name or ""

    ws[f"A{footer1_row}"] = f"审核人：{reviewer}  报账人：{reporter}  报账日期：{report_date}  合计金额：{total_amount}"
    ws[f"A{footer2_row}"] = f"收款人：{payee}  银行卡号：{bank_account}  开户行：{bank_name}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()